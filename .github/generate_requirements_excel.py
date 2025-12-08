#!/usr/bin/env python3
"""
Generate Excel spreadsheet from Requirements_Table.md
Converts markdown table format to Excel with section headers as a column for sorting
"""

import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


def parse_markdown_table(md_content):
    """Parse markdown content and extract table data with section headers."""
    
    rows = []
    current_section = None
    current_subsection = None
    current_category = None
    
    lines = md_content.split('\n')
    in_table = False
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
            
        # Detect main sections (## header)
        if line.startswith('## '):
            current_section = line.replace('##', '').strip()
            # Remove numbering like "1. INFRASTRUCTURE LAYER"
            current_section = re.sub(r'^\d+\.\s*', '', current_section)
            current_subsection = None
            current_category = None
            in_table = False
            continue
            
        # Detect subsections (### header)
        if line.startswith('### '):
            current_subsection = line.replace('###', '').strip()
            # Remove numbering like "1.1 Language Support & Editor"
            current_subsection = re.sub(r'^\d+\.\d+\s*', '', current_subsection)
            current_category = f"{current_section} > {current_subsection}"
            in_table = False
            continue
        
        # Detect table header (starts with | Req #)
        if line.startswith('| Req #'):
            in_table = True
            continue
            
        # Skip separator lines
        if in_table and re.match(r'^\|[\s\-:|]+\|$', line):
            continue
            
        # Parse table rows
        if in_table and line.startswith('|'):
            # Split by | and clean up
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            
            if len(cells) >= 5:
                req_id = cells[0]
                description = cells[1]
                spec = cells[2]
                phase = cells[3]
                status = cells[4]
                comments = cells[5] if len(cells) > 5 else ""
                
                # Clean up markdown formatting in cells
                description = re.sub(r'\*\*([^*]+)\*\*', r'\1', description)  # Remove **bold**
                comments = re.sub(r'\*\*([^*]+)\*\*', r'\1', comments)
                description = re.sub(r'`([^`]+)`', r'\1', description)  # Remove `code`
                comments = re.sub(r'`([^`]+)`', r'\1', comments)
                
                rows.append({
                    'Section': current_section or '',
                    'Subsection': current_subsection or '',
                    'Category': current_category or current_section or '',
                    'Req ID': req_id,
                    'Description': description,
                    'Spec': spec,
                    'Phase': phase,
                    'Status': status,
                    'Comments': comments
                })
        
        # Stop table parsing when we hit non-table content
        elif in_table and not line.startswith('|'):
            in_table = False
    
    return rows


def create_excel(rows, output_path):
    """Create Excel file from parsed data."""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Requirements"
    
    # Define headers
    headers = ['Section', 'Subsection', 'Category', 'Req ID', 'Description', 'Spec', 'Phase', 'Status', 'Comments']
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Status colors
    status_colors = {
        '✅': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),  # Green
        '🟡': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),  # Yellow
        '❌': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),  # Red
        '⚪': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'),  # Gray
    }
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows
    for row_num, row_data in enumerate(rows, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = row_data.get(header, '')
            cell.alignment = cell_alignment
            cell.border = border
            
            # Apply status color
            if header == 'Status':
                status_icon = cell.value.strip()
                if status_icon in status_colors:
                    cell.fill = status_colors[status_icon]
    
    # Set column widths
    column_widths = {
        'A': 30,  # Section
        'B': 35,  # Subsection
        'C': 50,  # Category
        'D': 12,  # Req ID
        'E': 60,  # Description
        'F': 12,  # Spec
        'G': 20,  # Phase
        'H': 8,   # Status
        'I': 80,  # Comments
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Enable auto-filter
    ws.auto_filter.ref = f"A1:I{len(rows) + 1}"
    
    # Save workbook
    wb.save(output_path)
    print(f"✅ Excel file created: {output_path}")
    print(f"   Total requirements: {len(rows)}")
    
    # Print statistics
    status_counts = {}
    for row in rows:
        status = row['Status'].strip()
        status_counts[status] = status_counts.get(status, 0) + 1
    
    print("\n📊 Status Summary:")
    for status, count in sorted(status_counts.items()):
        print(f"   {status}: {count}")


def main():
    """Main function to generate Excel from Requirements_Table.md"""
    
    # Paths
    script_dir = Path(__file__).parent
    md_file = script_dir / "Requirements_Table.md"
    excel_file = script_dir / "Requirements_Table.xlsx"
    
    # Check if markdown file exists
    if not md_file.exists():
        print(f"❌ Error: {md_file} not found!")
        sys.exit(1)
    
    # Read markdown file
    print(f"📖 Reading: {md_file}")
    with open(md_file, 'r', encoding='utf-8') as f:
        md_content = f.read()
    
    # Parse markdown
    print("🔍 Parsing markdown tables...")
    rows = parse_markdown_table(md_content)
    
    if not rows:
        print("❌ No table data found in markdown file!")
        sys.exit(1)
    
    # Create Excel
    print(f"📝 Creating Excel file...")
    create_excel(rows, excel_file)
    
    print(f"\n✅ Complete! Open the file to sort/filter by Section, Phase, Status, etc.")


if __name__ == '__main__':
    main()
