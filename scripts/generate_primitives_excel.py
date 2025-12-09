#!/usr/bin/env python3
"""
Generate comprehensive primitives status Excel sheet
Pillars: PERCEIVE, REASON, ACT, LEARN, Cross-Cutting
Implementation: Built-in (Rust/Tauri) or MCP
Status: ✅ IMPLEMENTED & INTEGRATED | 🟡 IMPLEMENTED NOT INTEGRATED | ❌ NOT IMPLEMENTED
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Define all primitives from Specifications.md Section 3.3
# Format: (Pillar, Category, Primitive, Protocol from Spec, Notes)
PRIMITIVES_DATA = [
    # PERCEIVE - Input & Sensing Layer
    ("PERCEIVE", "File Operations", "read_file", "Builtin", "Read file contents"),
    ("PERCEIVE", "File Operations", "list_files", "Builtin", "List files in directory (read_dir)"),
    ("PERCEIVE", "File Operations", "search_files", "Builtin", "Search for files by name/pattern (file_search)"),
    ("PERCEIVE", "File Operations", "file_metadata", "Builtin", "Get file size, modified time, permissions (get_file_info)"),
    
    ("PERCEIVE", "Dependency Analysis", "query_dependencies", "Builtin", "Find all dependencies of file (get_dependencies)"),
    ("PERCEIVE", "Dependency Analysis", "query_dependents", "Builtin", "Find all files depending on file (get_dependents)"),
    ("PERCEIVE", "Dependency Analysis", "find_imports", "Builtin", "Find all imports in file"),
    ("PERCEIVE", "Dependency Analysis", "find_callers", "Builtin", "Find all callers of function"),
    ("PERCEIVE", "Dependency Analysis", "impact_analysis", "Builtin", "Analyze impact of changing file"),
    ("PERCEIVE", "Dependency Analysis", "build_dependency_graph", "Builtin", "Generate full project graph (analyze_project)"),
    ("PERCEIVE", "Dependency Analysis", "get_module_boundaries", "Builtin", "Identify architectural layers"),
    
    ("PERCEIVE", "Code Intelligence", "parse_file", "Builtin", "Parse file into AST"),
    ("PERCEIVE", "Code Intelligence", "find_symbols", "LSP", "Find all functions/classes in file"),
    ("PERCEIVE", "Code Intelligence", "go_to_definition", "LSP", "Jump to symbol definition"),
    ("PERCEIVE", "Code Intelligence", "find_references", "LSP", "Find all uses of symbol"),
    ("PERCEIVE", "Code Intelligence", "hover_info", "LSP", "Get documentation for symbol"),
    
    ("PERCEIVE", "Test & Validation", "get_test_results", "Builtin", "Retrieve last test run results (execute_tests)"),
    ("PERCEIVE", "Test & Validation", "get_coverage", "Builtin", "Get code coverage metrics (execute_tests_with_coverage, get_test_coverage)"),
    ("PERCEIVE", "Test & Validation", "get_security_scan", "Builtin", "Get security scan results"),
    
    ("PERCEIVE", "Environment Sensing", "get_installed_packages", "Builtin", "List installed dependencies"),
    ("PERCEIVE", "Environment Sensing", "check_environment", "Builtin", "Verify Python/Node/Rust version (env_validate)"),
    ("PERCEIVE", "Environment Sensing", "get_git_status", "MCP", "Get Git repository status (git_status)"),
    ("PERCEIVE", "Environment Sensing", "env_get", "Builtin", "Get environment variable"),
    ("PERCEIVE", "Environment Sensing", "env_set", "Builtin", "Set environment variable"),
    
    ("PERCEIVE", "Browser Sensing", "get_console_logs", "Builtin (CDP)", "Get browser console output (browser_console_logs)"),
    ("PERCEIVE", "Browser Sensing", "get_network_logs", "Builtin (CDP)", "Get browser network requests"),
    ("PERCEIVE", "Browser Sensing", "capture_screenshot", "Builtin (CDP)", "Take browser screenshot (browser_screenshot)"),
    ("PERCEIVE", "Browser Sensing", "get_dom_element", "Builtin (CDP)", "Query DOM for element"),
    
    # REASON - Decision-Making & Analysis Layer
    ("REASON", "Pattern Matching", "search_codex_patterns", "Builtin", "Find similar code patterns (Yantra Codex - NOT IN TIER 1)"),
    ("REASON", "Pattern Matching", "find_bug_fix_patterns", "Builtin", "Find how similar bugs were fixed (Yantra Codex - NOT IN TIER 1)"),
    ("REASON", "Pattern Matching", "find_test_strategies", "Builtin", "Find successful test patterns (Yantra Codex - NOT IN TIER 1)"),
    ("REASON", "Pattern Matching", "find_api_patterns", "Builtin", "Find API design patterns used (Yantra Codex - NOT IN TIER 1)"),
    
    ("REASON", "Risk Assessment", "analyze_blast_radius", "Builtin", "Assess impact of proposed change (validate_dependencies, find_affected_tests)"),
    ("REASON", "Risk Assessment", "detect_breaking_changes", "Builtin", "Identify API breaking changes"),
    ("REASON", "Risk Assessment", "assess_test_coverage", "Builtin", "Check if change is well-tested (get_test_coverage)"),
    ("REASON", "Risk Assessment", "check_security_risk", "Builtin", "Evaluate security implications"),
    
    ("REASON", "Architectural Analysis", "check_architecture_alignment", "Builtin", "Verify code follows architecture"),
    ("REASON", "Architectural Analysis", "detect_boundary_violations", "Builtin", "Find component boundary violations"),
    ("REASON", "Architectural Analysis", "analyze_dependencies", "Builtin", "Check for circular dependencies (validate_dependencies)"),
    ("REASON", "Architectural Analysis", "validate_design_patterns", "Builtin", "Verify design patterns used correctly"),
    
    ("REASON", "LLM Consultation", "consult_primary_llm", "Builtin", "Get response from Claude Sonnet 4 (via LLMOrchestrator)"),
    ("REASON", "LLM Consultation", "consult_secondary_llm", "Builtin", "Get response from GPT-4 Turbo (via LLMOrchestrator)"),
    ("REASON", "LLM Consultation", "consult_specialist_llm", "Builtin", "Get response from specialized model (13 providers in multi_llm_manager.rs)"),
    ("REASON", "LLM Consultation", "aggregate_llm_responses", "Builtin", "Combine multiple LLM responses (multi_llm_manager.rs)"),
    
    # ACT - Execution & Action Layer
    ("ACT", "Code Generation", "generate_code", "Builtin", "Generate new code from spec (generate_code command)"),
    ("ACT", "Code Generation", "generate_tests", "Builtin", "Generate tests for code (generate_tests command)"),
    ("ACT", "Code Generation", "generate_documentation", "Builtin", "Generate API documentation"),
    ("ACT", "Code Generation", "refactor_code", "Builtin", "Refactor existing code"),
    
    ("ACT", "File Manipulation", "create_file", "Builtin", "Create new file with content (write_file)"),
    ("ACT", "File Manipulation", "modify_file", "Builtin", "Modify existing file (edit_file)"),
    ("ACT", "File Manipulation", "delete_file", "Builtin", "Delete file (file_delete)"),
    ("ACT", "File Manipulation", "move_file", "Builtin", "Move/rename file (file_move)"),
    
    ("ACT", "Test Execution", "run_tests", "Builtin", "Execute test suite (execute_tests)"),
    ("ACT", "Test Execution", "run_single_test", "Builtin", "Execute specific test"),
    ("ACT", "Test Execution", "run_coverage", "Builtin", "Execute tests with coverage (execute_tests_with_coverage)"),
    ("ACT", "Test Execution", "run_stress_tests", "Builtin", "Execute concurrency stress tests"),
    
    ("ACT", "Deployment", "deploy_local", "Builtin", "Deploy to localhost"),
    ("ACT", "Deployment", "deploy_railway", "MCP", "Deploy to Railway"),
    ("ACT", "Deployment", "deploy_aws", "MCP", "Deploy to AWS"),
    ("ACT", "Deployment", "health_check", "Builtin", "Check deployment health (api_health_check)"),
    ("ACT", "Deployment", "rollback_deployment", "Builtin", "Rollback to previous version"),
    
    ("ACT", "Browser Automation", "browser_navigate", "Builtin (CDP)", "Navigate to URL (browser_navigate)"),
    ("ACT", "Browser Automation", "browser_click", "Builtin (CDP)", "Click element (browser_click)"),
    ("ACT", "Browser Automation", "browser_fill_form", "Builtin (CDP)", "Fill form fields (browser_type)"),
    ("ACT", "Browser Automation", "browser_submit", "Builtin (CDP)", "Submit form"),
    ("ACT", "Browser Automation", "browser_wait", "Builtin (CDP)", "Wait for element"),
    
        ("ACT", "Git Operations", "git_setup", "Builtin", "Chat-based Git configuration & auth (NEW)"),
    ("ACT", "Git Operations", "git_authenticate", "Builtin", "Store credentials securely (NEW) - uses system keychain"),
    ("ACT", "Git Operations", "git_test_connection", "Builtin", "Validate authentication works (NEW)"),
    ("ACT", "Git Operations", "git_status", "MCP/Builtin", "Current state (via terminal commands)"),
    ("ACT", "Git Operations", "git_diff", "MCP/Builtin", "Changes (staged, unstaged, between refs)"),
    ("ACT", "Git Operations", "git_log", "MCP/Builtin", "Commit history"),
    ("ACT", "Git Operations", "git_blame", "MCP/Builtin", "Line-by-line attribution"),
    ("ACT", "Git Operations", "git_commit", "MCP/Builtin", "Create commit with auto-messages"),
    ("ACT", "Git Operations", "git_push", "MCP/Builtin", "Push commits to remote"),
    ("ACT", "Git Operations", "git_pull", "MCP/Builtin", "Pull latest changes"),
    ("ACT", "Git Operations", "git_branch", "MCP/Builtin", "Create/switch/list branches"),
    ("ACT", "Git Operations", "git_checkout", "MCP/Builtin", "Checkout files/branches"),
    ("ACT", "Git Operations", "git_merge", "MCP/Builtin", "Merge branches"),
    ("ACT", "Git Operations", "git_stash", "MCP/Builtin", "Stash/pop changes"),
    ("ACT", "Git Operations", "git_reset", "MCP/Builtin", "Undo changes"),
    ("ACT", "Git Operations", "git_clone", "MCP/Builtin", "Clone repository"),
    ("ACT", "Git Operations", "git_resolve_conflict", "Builtin", "AI-powered conflict resolution (Post-MVP)"),
    
    ("ACT", "YDoc Operations", "create_ydoc_document", "Builtin", "Create new YDoc document"),
    ("ACT", "YDoc Operations", "create_ydoc_block", "Builtin", "Create new block in document"),
    ("ACT", "YDoc Operations", "update_ydoc_block", "Builtin", "Update existing block"),
    ("ACT", "YDoc Operations", "link_ydoc_to_code", "Builtin", "Create graph edge doc → code"),
    ("ACT", "YDoc Operations", "search_ydoc_blocks", "Builtin", "Search documentation blocks"),
    
    ("ACT", "Terminal & Shell", "execute_command", "Builtin", "Execute shell command with streaming (execute_terminal_command)"),
    ("ACT", "Terminal & Shell", "manage_environment_vars", "Builtin", "Environment variable management (env_get, env_set)"),
    ("ACT", "Terminal & Shell", "control_working_directory", "Builtin", "Working directory control"),
    ("ACT", "Terminal & Shell", "manage_background_processes", "Builtin", "Background process management (execute_background)"),
    ("ACT", "Terminal & Shell", "capture_exit_codes", "Builtin", "Exit code capture and error handling"),
    
    # LEARN - Feedback & Adaptation Layer
    ("LEARN", "Pattern Capture", "record_success_pattern", "Builtin", "Store successful code pattern (Yantra Codex - NOT IN TIER 1)"),
    ("LEARN", "Pattern Capture", "record_bug_fix", "Builtin", "Store bug and fix pattern (Yantra Codex - NOT IN TIER 1)"),
    ("LEARN", "Pattern Capture", "record_test_strategy", "Builtin", "Store effective test pattern (Yantra Codex - NOT IN TIER 1)"),
    ("LEARN", "Pattern Capture", "record_llm_mistake", "Builtin", "Store LLM error and correction (Yantra Codex - NOT IN TIER 1)"),
    
    ("LEARN", "Feedback Processing", "process_test_failure", "Builtin", "Analyze why test failed"),
    ("LEARN", "Feedback Processing", "process_security_finding", "Builtin", "Analyze security vulnerability"),
    ("LEARN", "Feedback Processing", "process_user_feedback", "Builtin", "Incorporate user corrections"),
    ("LEARN", "Feedback Processing", "process_deployment_failure", "Builtin", "Analyze deployment issue"),
    
    ("LEARN", "Codex Updates", "update_pattern_confidence", "Builtin", "Adjust pattern confidence score (Yantra Codex - NOT IN TIER 1)"),
    ("LEARN", "Codex Updates", "increment_success_count", "Builtin", "Increment pattern success counter (Yantra Codex - NOT IN TIER 1)"),
    ("LEARN", "Codex Updates", "increment_failure_count", "Builtin", "Increment pattern failure counter (Yantra Codex - NOT IN TIER 1)"),
    ("LEARN", "Codex Updates", "archive_low_confidence_patterns", "Builtin", "Archive patterns below threshold (Yantra Codex - NOT IN TIER 1)"),
    
    ("LEARN", "Analytics", "track_generation_time", "Builtin", "Record how long code gen took"),
    ("LEARN", "Analytics", "track_test_pass_rate", "Builtin", "Record test success rate"),
    ("LEARN", "Analytics", "track_security_scan_results", "Builtin", "Record security findings"),
    ("LEARN", "Analytics", "track_deployment_success", "Builtin", "Record deployment outcomes"),
    
    # Cross-Cutting Primitives
    ("Cross-Cutting", "State Management", "save_agent_state", "Builtin", "Persist current agent state (AgentStateManager)"),
    ("Cross-Cutting", "State Management", "load_agent_state", "Builtin", "Restore agent state (AgentStateManager)"),
    ("Cross-Cutting", "State Management", "create_checkpoint", "Builtin", "Create rollback point"),
    ("Cross-Cutting", "State Management", "rollback_to_checkpoint", "Builtin", "Restore from checkpoint"),
    
    ("Cross-Cutting", "Context Management", "assemble_context", "Builtin", "Build context for LLM (assemble_hierarchical_context in orchestrator)"),
    ("Cross-Cutting", "Context Management", "compress_context", "Builtin", "Compress context to fit window"),
    ("Cross-Cutting", "Context Management", "prioritize_context", "Builtin", "Rank context by relevance"),
    ("Cross-Cutting", "Context Management", "cache_context", "Builtin", "Cache frequently used context"),
    ("Cross-Cutting", "Context Management", "context_add", "Builtin", "Add to persistent conversation DB"),
    ("Cross-Cutting", "Context Management", "context_search", "Builtin", "Search code + conversation"),
    ("Cross-Cutting", "Context Management", "context_summarize", "Builtin", "Compress code + conversation"),
    ("Cross-Cutting", "Context Management", "conversation_search", "Builtin", "Search past conversations (keyword/semantic)"),
    ("Cross-Cutting", "Context Management", "conversation_history", "Builtin", "Retrieve conversation for context"),
    ("Cross-Cutting", "Context Management", "conversation_link", "Builtin", "Link message to work session"),
    
    ("Cross-Cutting", "Communication", "send_user_message", "Builtin", "Send message to user"),
    ("Cross-Cutting", "Communication", "request_user_approval", "Builtin", "Request user decision"),
    ("Cross-Cutting", "Communication", "show_progress", "Builtin", "Update progress UI (status_update_progress)"),
    ("Cross-Cutting", "Communication", "log_event", "Builtin", "Log event for debugging"),
    
    ("Cross-Cutting", "Error Handling", "retry_with_backoff", "Builtin", "Retry failed operation (orchestrator auto-retry loop)"),
    ("Cross-Cutting", "Error Handling", "fallback_to_secondary", "Builtin", "Use backup strategy (multi_llm_manager failover)"),
    ("Cross-Cutting", "Error Handling", "report_error", "Builtin", "Report unrecoverable error"),
    ("Cross-Cutting", "Error Handling", "request_human_help", "Builtin", "Escalate to user (OrchestrationResult::Escalated)"),
]

# Map spec primitives to actual Tauri commands
# Format: {spec_name: (tauri_command, integration_status, file_location)}
COMMAND_MAPPINGS = {
    # PERCEIVE - File Operations
    "read_file": ("read_file", "✅", "main.rs:54"),
    "list_files": ("read_dir", "✅", "main.rs:86"),
    "search_files": ("file_search", "✅", "main.rs:632 + agent/file_ops.rs"),
    "file_metadata": ("get_file_info", "✅", "main.rs:130"),
    
    # PERCEIVE - Dependency Analysis
    "query_dependencies": ("get_dependencies", "✅", "main.rs:170"),
    "query_dependents": ("get_dependents", "✅", "main.rs:181"),
    "find_imports": ("parse_file", "🟡", "gnn/parser.rs (tree-sitter)"),
    "find_callers": ("find_node", "🟡", "main.rs:192 + gnn traversal"),
    "impact_analysis": ("validate_dependencies", "✅", "main.rs:734 + agent/validation.rs"),
    "build_dependency_graph": ("analyze_project", "✅", "main.rs:153"),
    "get_module_boundaries": ("N/A", "❌", "Not implemented"),
    
    # PERCEIVE - Code Intelligence (LSP)
    "parse_file": ("N/A", "🟡", "gnn/parser.rs (tree-sitter, not LSP)"),
    "find_symbols": ("N/A", "❌", "LSP not integrated"),
    "go_to_definition": ("N/A", "❌", "LSP not integrated"),
    "find_references": ("N/A", "❌", "LSP not integrated"),
    "hover_info": ("N/A", "❌", "LSP not integrated"),
    
    # PERCEIVE - Test & Validation
    "get_test_results": ("execute_tests", "✅", "main.rs:1372 + testing/runners/"),
    "get_coverage": ("execute_tests_with_coverage + get_test_coverage", "✅", "main.rs:1386,1211"),
    "get_security_scan": ("N/A", "❌", "Semgrep integration TODO"),
    
    # PERCEIVE - Environment
    "get_installed_packages": ("N/A", "🟡", "dependency_manager.rs (detect only)"),
    "check_environment": ("env_validate", "✅", "main.rs:969 + agent/environment.rs"),
    "get_git_status": ("git_status", "✅", "main.rs:1402 + git/manager.rs"),
    "env_get": ("N/A", "🟡", "agent/environment.rs (no command)"),
    "env_set": ("N/A", "🟡", "agent/environment.rs (no command)"),
    
    # PERCEIVE - Browser
    "get_console_logs": ("browser_console_logs", "✅", "main.rs:596 + browser/cdp.rs"),
    "get_network_logs": ("N/A", "❌", "CDP integration partial"),
    "capture_screenshot": ("browser_screenshot", "✅", "main.rs:578"),
    "get_dom_element": ("browser_evaluate_js", "🟡", "main.rs:587 (via JS eval)"),
    
    # REASON - Pattern Matching (Yantra Codex - NOT IN TIER 1)
    "search_codex_patterns": ("N/A", "❌", "Yantra Codex not in Tier 1"),
    "find_bug_fix_patterns": ("N/A", "❌", "Yantra Codex not in Tier 1"),
    "find_test_strategies": ("N/A", "❌", "Yantra Codex not in Tier 1"),
    "find_api_patterns": ("N/A", "❌", "Yantra Codex not in Tier 1"),
    
    # REASON - Risk Assessment
    "analyze_blast_radius": ("validate_dependencies + find_affected_tests", "✅", "main.rs:734,1062"),
    "detect_breaking_changes": ("N/A", "❌", "Not implemented"),
    "assess_test_coverage": ("get_test_coverage", "✅", "main.rs:1211"),
    "check_security_risk": ("N/A", "❌", "Security scanning TODO"),
    
    # REASON - Architectural Analysis
    "check_architecture_alignment": ("N/A", "❌", "Not implemented"),
    "detect_boundary_violations": ("N/A", "❌", "Not implemented"),
    "analyze_dependencies": ("validate_dependencies", "✅", "main.rs:734 + agent/validation.rs"),
    "validate_design_patterns": ("N/A", "❌", "Not implemented"),
    
    # REASON - LLM Consultation
    "consult_primary_llm": ("generate_code (via LLMOrchestrator)", "✅", "llm/orchestrator.rs"),
    "consult_secondary_llm": ("generate_code (via LLMOrchestrator)", "✅", "llm/orchestrator.rs:fallback"),
    "consult_specialist_llm": ("N/A", "🟡", "llm/multi_llm_manager.rs (13 providers)"),
    "aggregate_llm_responses": ("N/A", "🟡", "llm/multi_llm_manager.rs (logic exists)"),
    
    # ACT - Code Generation
    "generate_code": ("generate_code", "✅", "main.rs:1279 + agent/orchestrator.rs"),
    "generate_tests": ("generate_tests", "✅", "main.rs:1335 + testing/generators/"),
    "generate_documentation": ("N/A", "❌", "Not implemented"),
    "refactor_code": ("N/A", "❌", "Not implemented"),
    
    # ACT - File Manipulation
    "create_file": ("write_file", "✅", "main.rs:61"),
    "modify_file": ("edit_file", "✅", "main.rs:72 + agent/file_editor.rs"),
    "delete_file": ("file_delete", "✅", "main.rs:614 + agent/file_ops.rs"),
    "move_file": ("file_move", "✅", "main.rs:620 + agent/file_ops.rs"),
    
    # ACT - Test Execution
    "run_tests": ("execute_tests", "✅", "main.rs:1372 + testing/runners/"),
    "run_single_test": ("execute_tests", "🟡", "main.rs:1372 (supports specific test)"),
    "run_coverage": ("execute_tests_with_coverage", "✅", "main.rs:1386"),
    "run_stress_tests": ("N/A", "❌", "Not implemented"),
    
    # ACT - Deployment
    "deploy_local": ("N/A", "❌", "Not implemented"),
    "deploy_railway": ("N/A", "❌", "Railway MCP not integrated (MVP target)"),
    "deploy_aws": ("N/A", "❌", "AWS deployment not in MVP"),
    "health_check": ("api_health_check", "✅", "main.rs:914 + agent/api_health.rs"),
    "rollback_deployment": ("N/A", "❌", "Not implemented"),
    
    # ACT - Browser Automation
    "browser_navigate": ("browser_navigate", "✅", "main.rs:550 + browser/cdp.rs"),
    "browser_click": ("browser_click", "✅", "main.rs:560"),
    "browser_fill_form": ("browser_type", "✅", "main.rs:569"),
    "browser_submit": ("browser_click", "🟡", "main.rs:560 (can click submit)"),
    "browser_wait": ("N/A", "❌", "CDP wait not implemented"),
    
    # ACT - Git Operations (MCP protocol via git/mcp.rs)
    "git_commit": ("git_commit", "✅", "main.rs:1417 + git/mcp.rs (MCP protocol)"),
    "git_push": ("git_push", "✅", "main.rs:1468 + git/mcp.rs (MCP protocol)"),
    "git_branch": ("git_checkout", "🟡", "main.rs:1454 (checkout, not create branch)"),
    "git_merge": ("N/A", "❌", "Not implemented"),
    
    # ACT - YDoc Operations
    "create_ydoc_document": ("N/A", "❌", "YDoc system partial"),
    "create_ydoc_block": ("N/A", "❌", "YDoc system partial"),
    "update_ydoc_block": ("N/A", "❌", "YDoc system partial"),
    "link_ydoc_to_code": ("N/A", "❌", "YDoc system partial"),
    "search_ydoc_blocks": ("N/A", "❌", "YDoc system partial"),
    
    # ACT - Terminal
    "execute_command": ("execute_terminal_command", "✅", "main.rs:1547 + terminal/executor.rs"),
    "manage_environment_vars": ("env_create_snapshot", "🟡", "main.rs:946 (snapshot only)"),
    "control_working_directory": ("execute_terminal_command", "✅", "main.rs:1547 (supports cwd)"),
    "manage_background_processes": ("execute_background", "✅", "main.rs:790"),
    "capture_exit_codes": ("execute_terminal_command", "✅", "main.rs:1547 (returns exit code)"),
    
    # LEARN - Pattern Capture (Yantra Codex - NOT IN TIER 1)
    "record_success_pattern": ("N/A", "❌", "Yantra Codex not in Tier 1"),
    "record_bug_fix": ("N/A", "❌", "Yantra Codex not in Tier 1"),
    "record_test_strategy": ("N/A", "❌", "Yantra Codex not in Tier 1"),
    "record_llm_mistake": ("N/A", "❌", "Yantra Codex not in Tier 1"),
    
    # LEARN - Feedback Processing
    "process_test_failure": ("N/A", "🟡", "orchestrator.rs (auto-retry on test fail)"),
    "process_security_finding": ("N/A", "❌", "Not implemented"),
    "process_user_feedback": ("N/A", "❌", "Not implemented"),
    "process_deployment_failure": ("N/A", "❌", "Not implemented"),
    
    # LEARN - Codex Updates (Yantra Codex - NOT IN TIER 1)
    "update_pattern_confidence": ("N/A", "❌", "Yantra Codex not in Tier 1"),
    "increment_success_count": ("N/A", "❌", "Yantra Codex not in Tier 1"),
    "increment_failure_count": ("N/A", "❌", "Yantra Codex not in Tier 1"),
    "archive_low_confidence_patterns": ("N/A", "❌", "Yantra Codex not in Tier 1"),
    
    # LEARN - Analytics
    "track_generation_time": ("N/A", "🟡", "orchestrator.rs (tracks internally)"),
    "track_test_pass_rate": ("N/A", "🟡", "confidence.rs (tracks internally)"),
    "track_security_scan_results": ("N/A", "❌", "Not implemented"),
    "track_deployment_success": ("N/A", "❌", "Not implemented"),
    
    # Cross-Cutting - State Management
    "save_agent_state": ("N/A", "✅", "agent/state.rs:AgentStateManager::save_state"),
    "load_agent_state": ("N/A", "✅", "agent/state.rs:AgentStateManager::load_state"),
    "create_checkpoint": ("N/A", "❌", "Not implemented"),
    "rollback_to_checkpoint": ("N/A", "❌", "Not implemented"),
    
    # Cross-Cutting - Context Management
    "assemble_context": ("N/A", "✅", "llm/context.rs:assemble_hierarchical_context (used in orchestrator)"),
    "compress_context": ("N/A", "🟡", "llm/context.rs (logic exists)"),
    "prioritize_context": ("N/A", "🟡", "llm/context.rs (hierarchical L1-L4)"),
    "cache_context": ("N/A", "❌", "Not implemented"),
    "context_add": ("N/A", "❌", "Conversation DB not implemented (CONV-001 to CONV-016)"),
    "context_search": ("N/A", "❌", "Conversation DB not implemented (CONV-005)"),
    "context_summarize": ("N/A", "❌", "Not implemented"),
    "conversation_search": ("N/A", "❌", "Conversation DB not implemented (CONV-005)"),
    "conversation_history": ("N/A", "❌", "Conversation DB not implemented (CONV-003)"),
    "conversation_link": ("N/A", "❌", "Conversation DB not implemented (CONV-012, CONV-013)"),
    
    # Cross-Cutting - Communication
    "send_user_message": ("N/A", "❌", "Not implemented"),
    "request_user_approval": ("N/A", "❌", "Not implemented"),
    "show_progress": ("status_update_progress", "✅", "main.rs:660 + agent/status_transparency.rs"),
    "log_event": ("N/A", "🟡", "Standard Rust logging (tracing crate)"),
    
    # Cross-Cutting - Error Handling
    "retry_with_backoff": ("N/A", "✅", "agent/orchestrator.rs:auto-retry loop (up to 3 attempts)"),
    "fallback_to_secondary": ("N/A", "✅", "llm/orchestrator.rs:fallback mechanism"),
    "report_error": ("N/A", "✅", "OrchestrationResult::Error"),
    "request_human_help": ("N/A", "✅", "OrchestrationResult::Escalated"),
}

def create_excel_workbook():
    """Create Excel workbook with comprehensive primitives status"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agentic Primitives Status"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    status_fills = {
        "✅": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Light green
        "🟡": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Light yellow
        "❌": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # Light red
    }
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Pillar", "Category", "Primitive", "Implementation Method", "Status", "Tauri Command", "File Location", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Data rows
    row_idx = 2
    for pillar, category, primitive, protocol, notes in PRIMITIVES_DATA:
        # Get mapping info
        mapping = COMMAND_MAPPINGS.get(primitive, ("N/A", "❌", "Not found"))
        tauri_command, status, file_location = mapping
        
        # Determine implementation method
        if "MCP" in protocol:
            # Git operations actually use MCP protocol (git/mcp.rs)
            impl_method = "MCP"
        elif "LSP" in protocol:
            impl_method = "LSP (Not Integrated)"
        elif "CDP" in protocol:
            impl_method = "Built-in (CDP)"
        else:
            impl_method = "Built-in"
        
        # Write row
        ws.cell(row=row_idx, column=1).value = pillar
        ws.cell(row=row_idx, column=2).value = category
        ws.cell(row=row_idx, column=3).value = primitive
        ws.cell(row=row_idx, column=4).value = impl_method
        
        status_cell = ws.cell(row=row_idx, column=5)
        status_cell.value = status
        if status in status_fills:
            status_cell.fill = status_fills[status]
        
        ws.cell(row=row_idx, column=6).value = tauri_command
        ws.cell(row=row_idx, column=7).value = file_location
        ws.cell(row=row_idx, column=8).value = notes
        
        # Apply borders and alignment
        for col_idx in range(1, 9):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        row_idx += 1
    
    # Add summary section
    row_idx += 2
    summary_row = row_idx
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    summary_cell = ws.cell(row=summary_row, column=1)
    summary_cell.value = "SUMMARY"
    summary_cell.font = Font(bold=True, size=14)
    summary_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    summary_cell.font = Font(bold=True, size=14, color="FFFFFF")
    summary_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Count status
    implemented = sum(1 for _, _, primitive, _, _ in PRIMITIVES_DATA if COMMAND_MAPPINGS.get(primitive, ("", "❌", ""))[1] == "✅")
    partial = sum(1 for _, _, primitive, _, _ in PRIMITIVES_DATA if COMMAND_MAPPINGS.get(primitive, ("", "❌", ""))[1] == "🟡")
    not_impl = sum(1 for _, _, primitive, _, _ in PRIMITIVES_DATA if COMMAND_MAPPINGS.get(primitive, ("", "❌", ""))[1] == "❌")
    total = len(PRIMITIVES_DATA)
    
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = "✅ IMPLEMENTED & INTEGRATED:"
    ws.cell(row=row_idx, column=2).value = f"{implemented} ({implemented*100//total}%)"
    ws.cell(row=row_idx, column=2).fill = status_fills["✅"]
    
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = "🟡 IMPLEMENTED NOT INTEGRATED:"
    ws.cell(row=row_idx, column=2).value = f"{partial} ({partial*100//total}%)"
    ws.cell(row=row_idx, column=2).fill = status_fills["🟡"]
    
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = "❌ NOT IMPLEMENTED:"
    ws.cell(row=row_idx, column=2).value = f"{not_impl} ({not_impl*100//total}%)"
    ws.cell(row=row_idx, column=2).fill = status_fills["❌"]
    
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = "TOTAL PRIMITIVES:"
    ws.cell(row=row_idx, column=1).font = Font(bold=True)
    ws.cell(row=row_idx, column=2).value = total
    ws.cell(row=row_idx, column=2).font = Font(bold=True)
    
    # Add legend
    row_idx += 2
    ws.cell(row=row_idx, column=1).value = "LEGEND:"
    ws.cell(row=row_idx, column=1).font = Font(bold=True)
    
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = "✅"
    ws.cell(row=row_idx, column=2).value = "Tauri command exists AND agent orchestrator can use it"
    
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = "🟡"
    ws.cell(row=row_idx, column=2).value = "Tauri command exists OR logic exists, but NOT integrated for agent use"
    
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = "❌"
    ws.cell(row=row_idx, column=2).value = "Not implemented at all"
    
    # Add notes section
    row_idx += 2
    ws.cell(row=row_idx, column=1).value = "NOTES:"
    ws.cell(row=row_idx, column=1).font = Font(bold=True)
    
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = "- Yantra Codex (GNN learning system) is NOT in Tier 1 MVP (stretch goal), so all pattern/learning primitives are ❌"
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = "- LSP integration not implemented, all LSP primitives are ❌"
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = "- Git operations use MCP protocol (git/mcp.rs) as specified"
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = "- Conversation Memory System (CONV-001 to CONV-016) not implemented yet - database schema defined but no code"
    row_idx += 1
    ws.cell(row=row_idx, column=1).value = "- Railway deployment is MVP target, AWS is not in MVP scope"
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15  # Pillar
    ws.column_dimensions['B'].width = 22  # Category
    ws.column_dimensions['C'].width = 30  # Primitive
    ws.column_dimensions['D'].width = 20  # Implementation Method
    ws.column_dimensions['E'].width = 10  # Status
    ws.column_dimensions['F'].width = 35  # Tauri Command
    ws.column_dimensions['G'].width = 30  # File Location
    ws.column_dimensions['H'].width = 50  # Notes
    
    # Freeze panes (header row)
    ws.freeze_panes = 'A2'
    
    return wb

def main():
    """Generate Excel file"""
    print("🚀 Generating Agentic Primitives Status Excel sheet...")
    
    wb = create_excel_workbook()
    
    filename = f"Agentic_Primitives_Status_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    print(f"✅ Excel file generated: {filename}")
    print(f"\n📊 Statistics:")
    
    implemented = sum(1 for _, _, primitive, _, _ in PRIMITIVES_DATA if COMMAND_MAPPINGS.get(primitive, ("", "❌", ""))[1] == "✅")
    partial = sum(1 for _, _, primitive, _, _ in PRIMITIVES_DATA if COMMAND_MAPPINGS.get(primitive, ("", "❌", ""))[1] == "🟡")
    not_impl = sum(1 for _, _, primitive, _, _ in PRIMITIVES_DATA if COMMAND_MAPPINGS.get(primitive, ("", "❌", ""))[1] == "❌")
    total = len(PRIMITIVES_DATA)
    
    print(f"✅ IMPLEMENTED & INTEGRATED: {implemented} ({implemented*100//total}%)")
    print(f"🟡 IMPLEMENTED NOT INTEGRATED: {partial} ({partial*100//total}%)")
    print(f"❌ NOT IMPLEMENTED: {not_impl} ({not_impl*100//total}%)")
    print(f"📦 TOTAL PRIMITIVES: {total}")
    
    # Breakdown by pillar
    print(f"\n📋 Breakdown by Pillar:")
    for pillar in ["PERCEIVE", "REASON", "ACT", "LEARN", "Cross-Cutting"]:
        pillar_primitives = [(p, c, pr, pm, n) for p, c, pr, pm, n in PRIMITIVES_DATA if p == pillar]
        pillar_impl = sum(1 for _, _, pr, _, _ in pillar_primitives if COMMAND_MAPPINGS.get(pr, ("", "❌", ""))[1] == "✅")
        pillar_total = len(pillar_primitives)
        print(f"   {pillar}: {pillar_impl}/{pillar_total} ({pillar_impl*100//pillar_total if pillar_total > 0 else 0}%)")

if __name__ == "__main__":
    main()
